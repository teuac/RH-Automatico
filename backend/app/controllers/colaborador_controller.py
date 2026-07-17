import io
import json
import pandas as pd
from fastapi import Depends, HTTPException, status, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from app.database.session import get_db
from app.repositories.colaborador import colaborador_repository
from app.repositories.obra import obra_repository
from app.repositories.audit import audit_repository
from app.schemas.colaborador import (
    ColaboradorCreate, ColaboradorUpdate, ColaboradorResponse,
    ColaboradorImportResponse, MigrarObraRequest
)
from app.auth.rbac import get_active_user
from app.models.user import User


class ColaboradorController:

    # ------------------------------------------------------------------ #
    #  Listagem                                                            #
    # ------------------------------------------------------------------ #
    @staticmethod
    def get_all(
        skip: int = 0,
        limit: int = 200,
        obra_id: Optional[int] = None,
        search: Optional[str] = None,
        status: Optional[str] = None,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_active_user)
    ):
        return colaborador_repository.get_multi(db, skip, limit, obra_id, search, status)

    # ------------------------------------------------------------------ #
    #  Criar                                                               #
    # ------------------------------------------------------------------ #
    @staticmethod
    def create(
        request: Request,
        payload: ColaboradorCreate,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_active_user)
    ) -> ColaboradorResponse:
        # Verifica duplicidade de matrícula
        existing = colaborador_repository.get_by_matricula(db, payload.matricula)
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Matrícula '{payload.matricula}' já cadastrada."
            )
        # Verifica se obra existe
        if payload.obra_id:
            obra = obra_repository.get(db, payload.obra_id)
            if not obra:
                raise HTTPException(status_code=404, detail="Obra não encontrada.")

        colaborador = colaborador_repository.create(db, payload.model_dump())

        audit_repository.log(
            db=db, user_id=current_user.id, user_name=current_user.full_name,
            user_email=current_user.email,
            ip_address=request.client.host if request.client else "unknown",
            user_agent=request.headers.get("user-agent", "unknown"),
            module="Colaboradores", screen="Colaboradores", action="CREATE",
            description=f"Cadastrou colaborador: {colaborador.nome} ({colaborador.matricula})",
            object_changed="colaboradores", object_id=str(colaborador.id),
            result="SUCESSO", after_state=json.dumps(payload.model_dump())
        )
        return colaborador

    # ------------------------------------------------------------------ #
    #  Atualizar                                                           #
    # ------------------------------------------------------------------ #
    @staticmethod
    def update(
        colaborador_id: int,
        request: Request,
        payload: ColaboradorUpdate,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_active_user)
    ) -> ColaboradorResponse:
        colab = colaborador_repository.get(db, colaborador_id)
        if not colab:
            raise HTTPException(status_code=404, detail="Colaborador não encontrado.")

        # Checa duplicidade de matrícula se estiver alterando
        updated_data = payload.model_dump(exclude_unset=True)
        if "matricula" in updated_data and updated_data["matricula"] != colab.matricula:
            if colaborador_repository.get_by_matricula(db, updated_data["matricula"]):
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"Matrícula '{updated_data['matricula']}' já pertence a outro colaborador."
                )

        before = {"nome": colab.nome, "matricula": colab.matricula, "funcao": colab.funcao}
        updated = colaborador_repository.update(db, colab, updated_data)

        audit_repository.log(
            db=db, user_id=current_user.id, user_name=current_user.full_name,
            user_email=current_user.email,
            ip_address=request.client.host if request.client else "unknown",
            user_agent=request.headers.get("user-agent", "unknown"),
            module="Colaboradores", screen="Colaboradores", action="UPDATE",
            description=f"Atualizou colaborador: {updated.nome} ({updated.matricula})",
            object_changed="colaboradores", object_id=str(colaborador_id),
            result="SUCESSO", before_state=json.dumps(before), after_state=json.dumps(updated_data)
        )
        return updated

    # ------------------------------------------------------------------ #
    #  Deletar                                                             #
    # ------------------------------------------------------------------ #
    @staticmethod
    def delete(
        colaborador_id: int,
        request: Request,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_active_user)
    ):
        colab = colaborador_repository.get(db, colaborador_id)
        if not colab:
            raise HTTPException(status_code=404, detail="Colaborador não encontrado.")

        before = {"nome": colab.nome, "matricula": colab.matricula}
        colaborador_repository.remove(db, colaborador_id)

        audit_repository.log(
            db=db, user_id=current_user.id, user_name=current_user.full_name,
            user_email=current_user.email,
            ip_address=request.client.host if request.client else "unknown",
            user_agent=request.headers.get("user-agent", "unknown"),
            module="Colaboradores", screen="Colaboradores", action="DELETE",
            description=f"Removeu colaborador: {before['nome']} ({before['matricula']})",
            object_changed="colaboradores", object_id=str(colaborador_id),
            result="SUCESSO", before_state=json.dumps(before)
        )
        return {"detail": "Colaborador removido com sucesso."}

    # ------------------------------------------------------------------ #
    #  Download do Modelo Excel                                            #
    # ------------------------------------------------------------------ #
    @staticmethod
    def download_template(
        db: Session = Depends(get_db),
        current_user: User = Depends(get_active_user)
    ):
        wb = Workbook()
        ws = wb.active
        ws.title = "Colaboradores"

        # Estilos
        header_fill = PatternFill("solid", fgColor="1A73E8")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill("solid", fgColor="E8F0FE")
        example_font = Font(name="Calibri", color="3C4043", size=10)
        thin_border = Border(
            left=Side(style="thin", color="DADCE0"),
            right=Side(style="thin", color="DADCE0"),
            top=Side(style="thin", color="DADCE0"),
            bottom=Side(style="thin", color="DADCE0"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # Cabeçalho principal
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = "MODELO DE IMPORTAÇÃO — COLABORADORES"
        title_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
        title_cell.fill = PatternFill("solid", fgColor="0D47A1")
        title_cell.alignment = center_align

        # Subtítulo
        ws.merge_cells("A2:E2")
        sub_cell = ws["A2"]
        sub_cell.value = "Preencha os campos abaixo. Matrícula e Nome são obrigatórios. Código da Obra deve corresponder a uma obra ativa no sistema."
        sub_cell.font = Font(name="Calibri", italic=True, color="5F6368", size=9)
        sub_cell.fill = PatternFill("solid", fgColor="F1F3F4")
        sub_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 28

        # Cabeçalhos das colunas
        headers = ["Matrícula *", "Nome Completo *", "Função", "Código da Obra", "Status"]
        header_row = 3
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 22

        # Exemplos de preenchimento
        examples = [
            ["001234", "João da Silva Santos", "Pedreiro", "OBR-001", "ATIVO"],
            ["001235", "Maria Oliveira Costa", "Auxiliar de Serviços", "OBR-001", "ATIVO"],
            ["001236", "Carlos Eduardo Souza", "Encarregado", "OBR-002", "ATIVO"],
        ]
        for row_idx, row_data in enumerate(examples, header_row + 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = example_font
                cell.fill = example_fill
                cell.border = thin_border
                cell.alignment = left_align if col_idx == 2 else center_align
            ws.row_dimensions[row_idx].height = 18

        # Ajustar largura das colunas
        col_widths = [16, 35, 28, 20, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Legenda
        legend_row = header_row + len(examples) + 2
        ws.merge_cells(f"A{legend_row}:E{legend_row}")
        leg_cell = ws[f"A{legend_row}"]
        leg_cell.value = "* Campos obrigatórios   |   Status: ATIVO ou INATIVO   |   Linhas com matrícula duplicada serão ignoradas automaticamente"
        leg_cell.font = Font(name="Calibri", bold=False, color="5F6368", size=8, italic=True)
        leg_cell.alignment = left_align

        # Salvar em buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=modelo_importacao_colaboradores.xlsx"}
        )

    # ------------------------------------------------------------------ #
    #  Importação em lote                                                  #
    # ------------------------------------------------------------------ #
    @staticmethod
    async def import_excel(
        request: Request,
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        current_user: User = Depends(get_active_user)
    ) -> ColaboradorImportResponse:
        if not file.filename.endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Formato inválido. Envie um arquivo .xlsx ou .xls.")

        content = await file.read()
        try:
            df = pd.read_excel(io.BytesIO(content), header=2)  # linha 3 é o cabeçalho (0-indexed: 2)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Erro ao ler arquivo Excel: {str(e)}")

        # Normalizar colunas
        df.columns = [str(c).strip().lower() for c in df.columns]

        # Mapear colunas
        col_map = {}
        for col in df.columns:
            if "matr" in col:
                col_map["matricula"] = col
            elif "nome" in col:
                col_map["nome"] = col
            elif "fun" in col:
                col_map["funcao"] = col
            elif "obra" in col or "cod" in col:
                col_map["obra_codigo"] = col
            elif "status" in col:
                col_map["status"] = col

        if "matricula" not in col_map or "nome" not in col_map:
            raise HTTPException(
                status_code=422,
                detail="Arquivo não contém colunas de Matrícula e Nome. Use o modelo disponibilizado."
            )

        # Buscar obras para lookup por código
        from app.models.obra import Obra
        obras_map = {o.codigo: o.id for o in db.query(Obra).filter(Obra.status == "ATIVO").all()}

        items_to_create = []
        erros = []
        ignorados = 0

        for idx, row in df.iterrows():
            matricula = str(row.get(col_map.get("matricula", ""), "")).strip()
            nome = str(row.get(col_map.get("nome", ""), "")).strip()

            if not matricula or matricula in ("nan", "") or not nome or nome in ("nan", ""):
                continue  # linha vazia

            # Verifica duplicidade
            if colaborador_repository.get_by_matricula(db, matricula):
                ignorados += 1
                erros.append(f"Linha {idx + 4}: matrícula '{matricula}' já existe — ignorada.")
                continue

            funcao = str(row.get(col_map.get("funcao", ""), "")).strip()
            funcao = funcao if funcao not in ("nan", "") else None

            obra_codigo = str(row.get(col_map.get("obra_codigo", ""), "")).strip()
            obra_id = obras_map.get(obra_codigo) if obra_codigo not in ("nan", "", "None") else None
            if obra_codigo and obra_codigo not in ("nan", "") and obra_id is None:
                erros.append(f"Linha {idx + 4}: código de obra '{obra_codigo}' não encontrado — colaborador será cadastrado sem obra.")

            status_val = str(row.get(col_map.get("status", ""), "ATIVO")).strip().upper()
            if status_val not in ("ATIVO", "INATIVO"):
                status_val = "ATIVO"

            items_to_create.append({
                "matricula": matricula,
                "nome": nome,
                "funcao": funcao,
                "obra_id": obra_id,
                "status": status_val,
            })

        created = colaborador_repository.bulk_create(db, items_to_create)

        audit_repository.log(
            db=db, user_id=current_user.id, user_name=current_user.full_name,
            user_email=current_user.email,
            ip_address=request.client.host if request.client else "unknown",
            user_agent=request.headers.get("user-agent", "unknown"),
            module="Colaboradores", screen="Colaboradores", action="IMPORT",
            description=f"Importação em lote: {len(created)} criados, {ignorados} ignorados.",
            object_changed="colaboradores", object_id="bulk",
            result="SUCESSO"
        )

        return ColaboradorImportResponse(
            total=len(items_to_create) + ignorados,
            importados=len(created),
            ignorados=ignorados,
            erros=erros
        )

    # ------------------------------------------------------------------ #
    #  Migração de obra em massa                                           #
    # ------------------------------------------------------------------ #
    @staticmethod
    def migrar_obra(
        request: Request,
        payload: MigrarObraRequest,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_active_user)
    ):
        obra = obra_repository.get(db, payload.nova_obra_id)
        if not obra:
            raise HTTPException(status_code=404, detail="Obra destino não encontrada.")

        count = colaborador_repository.migrate_obra(db, payload.ids, payload.nova_obra_id)

        audit_repository.log(
            db=db, user_id=current_user.id, user_name=current_user.full_name,
            user_email=current_user.email,
            ip_address=request.client.host if request.client else "unknown",
            user_agent=request.headers.get("user-agent", "unknown"),
            module="Colaboradores", screen="Colaboradores", action="MIGRAR_OBRA",
            description=f"Migrou {count} colaborador(es) para obra '{obra.nome}' ({obra.codigo}).",
            object_changed="colaboradores", object_id="bulk",
            result="SUCESSO",
            after_state=json.dumps({"nova_obra_id": payload.nova_obra_id, "ids": payload.ids})
        )

        return {
            "detail": f"{count} colaborador(es) migrado(s) com sucesso para '{obra.nome}'.",
            "migrados": count,
            "obra_destino": obra.nome
        }


colaborador_controller = ColaboradorController()
